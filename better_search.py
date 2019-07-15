from openpyxl import load_workbook #initializes the module load_workbook

#gets the file you are searching must include .xlsx
file_name = input("Enter file name:  ")
sheet_name = input("Enter the sheet name:  ")

values_mode = input("Values mode?  ") #values mode returns the values in the row no values mode returns the row which something is on

#dictionaries for accepted values
yes = ['y','yes','YES','Yes','1']
no = ['no','n','0','No','NO']

#loading the file
wb = load_workbook(file_name)
ws = wb[sheet_name]

#non value mode function to return row 
def nVmode():
    while True:#search function to find the row in which an agreement exists
        search  = str(input("Look up something:  "))

        for row in ws.iter_rows(min_row = 1, max_row = 17147, min_col = 1, max_col = 16):
            for cell in row:
                if search == cell.value:
                    print('value found')
                    print(cell.row)
                    break
                else:
                    continue
                continue
#Value mode function to toggle value mode
def Vmode():
    while True:
        search = str(input("Look up something:  "))

        for row in ws.iter_rows(min_row = 1, max_row = 18000, min_col = 1, max_col = 16):
            for cell in row:
                if search == cell.value:
                    for cell in row:
                        print(cell.value)
                    break
                else:
                    continue
                continue

#the real program 
if values_mode in yes:
    Vmode()
elif values_mode in no:
    nVmode()
else:
    exit("you messed up")
